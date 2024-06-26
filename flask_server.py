from flask import Flask, render_template, request, jsonify
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows  # 确保包含这个导入
import os
import uuid

app = Flask(__name__)

EXCEL_FILE = 'ribao.xlsx'

@app.route('/')
def home():
    # 假设index.html位于templates文件夹内
    return render_template('index.html')

@app.route('/query_reports', methods=['GET'])
def query_reports():
    name = request.args.get('name')
    date = request.args.get('date')

    book = load_workbook(EXCEL_FILE)
    filtered_data = []

    for sheet_name in book.sheetnames:
        sheet = book[sheet_name]
        df = pd.DataFrame(sheet.values)
        if df.empty:
            continue

        # 假定姓名在第一列，日期在第二列
        headers = df.iloc[0]
        df = df[1:]
        df.columns = headers

        # 根据姓名和日期过滤
        # 如果不想查全部，就用这个代码
        # if name:
        #     df = df[df['姓名'] == name]
        if name == '请选择姓名':
            # 如果姓名是"请选择姓名",则不过滤姓名,返回所有人的数据
            pass
        elif name:
            df = df[df['姓名'] == name]
        if date:
            df = df[df['日期'] == date]

        if not df.empty:
            filtered_data.append((sheet_name, df.to_html(classes='table')))

    return jsonify(filtered_data)


@app.route('/delete_report', methods=['POST'])
def delete_report():
    data = request.json
    uuid = data['uuid']
    sheet_name = data['sheet_name']

    book = load_workbook(EXCEL_FILE)
    sheet = book[sheet_name]

    rows = list(sheet.iter_rows(min_row=2, values_only=True))  # 将生成器转换为列表

    row_to_delete = None
    for row in rows:
        if row[2] == uuid:
            row_to_delete = row
            break

    if row_to_delete is None:
        return jsonify({'message': '未找到匹配的 UUID'})

    row_idx = rows.index(row_to_delete) + 2  # 在列表上调用 index 方法

    sheet.delete_rows(row_idx)

    book.save(EXCEL_FILE)
    return jsonify({'message': '删除成功'})


@app.route('/submit_report', methods=['POST'])
def submit_report():
    # 解析JSON请求体
    data = request.json
    category = data['category']
    name = data['name']
    date = data['date']
    additional_fields = data['additionalFields']

    if not os.path.exists(EXCEL_FILE):
        pd.DataFrame().to_excel(EXCEL_FILE)
    book = load_workbook(EXCEL_FILE)

    if category not in book.sheetnames:
        book.create_sheet(title=category)
    sheet = book[category]

    # 生成 UUID
    unique_id = str(uuid.uuid4())

    # 创建一个DataFrame
    df = pd.DataFrame([{
        '姓名': name,
        '日期': date,
        'UUID': unique_id,  # 添加 UUID 列
        **additional_fields
    }])

    # 找到工作表中最后一行有数据的行号
    max_row = sheet.max_row
    if max_row == 1 and not list(sheet.rows)[0]:  # 检查是否有标题行
        max_row = 0  # 如果没有标题和数据，则从第一行开始

    # 使用dataframe_to_rows转换DataFrame，并从下一行开始插入数据
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=max_row + 1):
        for c_idx, value in enumerate(row, start=1):
            sheet.cell(row=r_idx, column=c_idx, value=value)

    book.save(EXCEL_FILE)
    return jsonify({'message': 'Report submitted successfully'})


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0')